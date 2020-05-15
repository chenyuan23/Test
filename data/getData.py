from openpyxl import load_workbook

wb = load_workbook(r'D:\pycharm\test\data\平台.xlsx')
ws = wb['Sheet1']
print(ws.max_row)
print(ws.sheet(1, 1).value)

all_data = []
for i in range(2, ws.max_row+1):
    datas = {}
    datas['case_id'] = ws.cell(i, 1).value

    all_data.append(datas)